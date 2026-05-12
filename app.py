import streamlit as st
import pandas as pd
import openpyxl
from datetime import datetime, timedelta
import re
import io

# --- PAGE CONFIG ---
st.set_page_config(page_title="RDM Timesheet Auditor", page_icon="📊")

def parse_date_manual(user_input):
    try:
        clean_input = re.split(r'[_ ]', user_input)
        day = int(clean_input[0])
        month_str = clean_input[1].capitalize()[:3]
        current_year = 2026 
        month_map = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
                     'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12}
        return datetime(current_year, month_map.get(month_str, 1), day)
    except:
        return None

def clean_name(name):
    if pd.isna(name): return ""
    return re.sub(r'[^a-zA-Z0-9]', '', str(name).upper())

def add_time_column(df):
    """
    Automatically add the time extraction column after 'Date Time'
    Mimics Excel formula: =TIMEVALUE(RIGHT(G2,5))
    """
    # Extract the last 5 characters (HH:MM) from Date Time and convert to time
    df['TimeExtracted'] = df['Date Time'].apply(lambda x: 
        pd.to_datetime(str(x)[-5:], format='%H:%M').time() if pd.notna(x) and len(str(x)) >= 5 else None
    )
    return df

def process_data(ts_file, att_file, start_date):
    # Date Setup
    end_date = start_date + timedelta(days=6)
    date_range = [(start_date + timedelta(days=i)).date() for i in range(7)]
    dynamic_sheet_name = f"{start_date.day} {start_date.strftime('%b')} - {end_date.day} {end_date.strftime('%b')}"

    # Load Files
    if ts_file.name.endswith('.csv'):
        df = pd.read_csv(ts_file)
    else:
        df = pd.read_excel(ts_file)
    
    # NEW: Automatically add the time column if it doesn't exist
    if 'TimeExtracted' not in df.columns and 'Unnamed: 7' not in df.columns:
        df = add_time_column(df)
        st.info("✓ Time column automatically added to timesheet data")
    elif 'Unnamed: 7' in df.columns:
        # File already has the time column
        df['TimeExtracted'] = df['Unnamed: 7']
    
    # 1. Precise DateTime Parsing
    if 'Date Time' in df.columns:
        df['DateTime'] = pd.to_datetime(
            df['Date Time'].str.extract(r'(\d{4}-\d{2}-\d{2})')[0] + ' ' + 
            df['Date Time'].str.extract(r'(\d{2}:\d{2})')[0]
        )
    
    # 2. Assign "Operational Day" (Shift Day)
    df['WorkDate'] = df.apply(lambda r: (r['DateTime'] - timedelta(hours=5)).date(), axis=1)
    df['TimeOnly'] = df['DateTime'].dt.time
    df['CleanName'] = df['Name'].apply(clean_name)

    try:
        df_attendance = pd.read_excel(att_file, sheet_name=dynamic_sheet_name, header=None)
    except:
        df_attendance = pd.read_excel(att_file, sheet_name=0, header=None)
    
    # Extract Names (Rows 2 to 100)
    raw_employee_list = [df_attendance.iloc[idx, 1] for idx in range(2, min(100, len(df_attendance))) 
                         if pd.notna(df_attendance.iloc[idx, 1])]
    attendance_map = {clean_name(name): name for name in raw_employee_list}
    
    output_data = []
    exception_logs = []
    remark_logs = []  # NEW: For remarks sheet

    for cleaned_name, original_name in attendance_map.items():
        row_times = []
        emp_records = df[df['CleanName'] == cleaned_name].sort_values('DateTime')

        for work_date in date_range:
            # Filter logs strictly for this "Operational Day"
            day_logs = emp_records[emp_records['WorkDate'] == work_date]
            
            login_time = None
            logout_time = None
            has_login = False
            has_logout = False
            
            if not day_logs.empty:
                # --- LOGIN LOGIC ---
                starts = day_logs[day_logs['Type'] == 'Start Work']
                sites_in = day_logs[day_logs['Type'] == 'Site In']
                
                if not starts.empty:
                    login_time = starts.sort_values('DateTime').iloc[0]['TimeOnly']
                    has_login = True
                elif not sites_in.empty:
                    login_time = sites_in.sort_values('DateTime').iloc[0]['TimeOnly']
                    has_login = True
                else:
                    login_time = "NO LOGIN"

                # --- LOGOUT LOGIC ---
                valid_logouts = day_logs[day_logs['Type'].isin(['End Work', 'Site Out'])]
                
                if not valid_logouts.empty:
                    last_logout_event = valid_logouts.sort_values('DateTime').iloc[-1]
                    logout_time = last_logout_event['TimeOnly']
                    has_logout = True
                else:
                    logout_time = "NO LOGOUT"

                # --- CLEANUP FOR EMPTY DAYS ---
                if login_time == "NO LOGIN" and logout_time == "NO LOGOUT":
                    login_time = None
                    logout_time = None

                # --- EXCEPTION LOGGING ---
                if logout_time == "NO LOGOUT" and login_time is not None:
                    exception_logs.append({'Name': original_name, 'Date': work_date, 'Time': login_time, 'Reason': 'Missing Logout'})
                elif login_time == "NO LOGIN" and logout_time is not None:
                    exception_logs.append({'Name': original_name, 'Date': work_date, 'Time': logout_time, 'Reason': 'Logout without Login'})

                # --- REMARK LOGGING (NEW) ---
                # Check if any records for this day have remarks
                day_remarks = day_logs[day_logs['Remark'].notna()]
                for _, remark_row in day_remarks.iterrows():
                    remark_logs.append({
                        'Name': original_name,
                        'Day': work_date.strftime('%d %b'),
                        'Time': remark_row['TimeOnly'].strftime('%H:%M') if pd.notna(remark_row['TimeOnly']) else '',
                        'Login': 'Yes' if has_login else 'No',
                        'Logout': 'Yes' if has_logout else 'No',
                        'Remark': remark_row['Remark']
                    })

            # Handle explicit None for output
            row_times.extend([login_time if login_time else None, 
                              logout_time if logout_time else None])
                              
        output_data.append(row_times)

    return pd.DataFrame(output_data), pd.DataFrame(exception_logs), pd.DataFrame(remark_logs), date_range, raw_employee_list

# --- UI DESIGN ---
st.title("📊 Timesheet Audit Web Portal")
st.markdown("Upload your files below to generate the audit report.")

col1, col2 = st.columns(2)

with col1:
    ts_file = st.file_uploader("Upload Timesheet Detail (Excel/CSV)", type=['xlsx', 'csv'])
with col2:
    att_file = st.file_uploader("Upload Attendance_New.xlsx", type=['xlsx'])

start_date_str = st.text_input("Enter Start Date (e.g., 23_Jan or 30_Jan)", "")

if st.button("🚀 Generate Audit Report"):
    if ts_file and att_file and start_date_str:
        start_date = parse_date_manual(start_date_str)
        if start_date:
            with st.spinner('Processing...'):
                df_out, df_ex, df_remarks, d_range, names = process_data(ts_file, att_file, start_date)
                
                # Show Preview
                st.success("✅ Analysis Complete!")
                
                col_prev1, col_prev2, col_prev3 = st.columns(3)
                with col_prev1:
                    st.metric("Employees Processed", len(names))
                with col_prev2:
                    st.metric("Exceptions Found", len(df_ex))
                with col_prev3:
                    st.metric("Remarks Found", len(df_remarks))
                
                st.subheader("📋 Summary Preview")
                st.dataframe(df_out.head(10))
                
                if len(df_remarks) > 0:
                    st.subheader("💬 Remarks Preview")
                    st.dataframe(df_remarks.head(10))
                
                # Prepare Excel with 3 sheets: Summary, Exceptions, Remarks
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    # --- SUMMARY SHEET ---
                    wb = writer.book
                    ws = wb.create_sheet("Summary")
                    
                    # Formatting
                    bold = openpyxl.styles.Font(bold=True)
                    center = openpyxl.styles.Alignment(horizontal='center')
                    red_text = openpyxl.styles.Font(color="FF0000", bold=True)

                    # Headers
                    ws.cell(1, 1, "No").font = bold
                    ws.cell(1, 2, "Employee Name").font = bold
                    
                    for i, d_obj in enumerate(d_range):
                        col = 3 + (i * 2)
                        ws.cell(1, col, d_obj.strftime('%d %b')).font = bold
                        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)
                        ws.cell(1, col).alignment = center
                        ws.cell(2, col, "Login").font = bold
                        ws.cell(2, col+1, "Logout").font = bold

                    # Data rows
                    for r_idx, (emp_name, row_vals) in enumerate(zip(names, df_out.itertuples(index=False)), start=3):
                        ws.cell(r_idx, 1, r_idx - 2)
                        ws.cell(r_idx, 2, emp_name)
                        for c_idx, val in enumerate(row_vals, start=3):
                            cell = ws.cell(r_idx, c_idx, val)
                            
                            # Handle "NO LOGOUT" / "NO LOGIN" red text
                            if val in ["NO LOGOUT", "NO LOGIN"]:
                                cell.font = red_text
                                cell.value = val
                            elif val is not None:
                                cell.value = val
                                cell.number_format = 'HH:MM'
                    
                    # --- EXCEPTIONS SHEET ---
                    df_ex.to_excel(writer, index=False, sheet_name="Exceptions")
                    
                    # --- REMARKS SHEET (NEW) ---
                    if len(df_remarks) > 0:
                        df_remarks.to_excel(writer, index=False, sheet_name="Remarks")
                        
                        # Format the Remarks sheet
                        ws_remarks = wb["Remarks"]
                        
                        # Bold headers
                        for cell in ws_remarks[1]:
                            cell.font = bold
                        
                        # Auto-adjust column widths
                        for column in ws_remarks.columns:
                            max_length = 0
                            column_letter = openpyxl.utils.get_column_letter(column[0].column)
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            ws_remarks.column_dimensions[column_letter].width = adjusted_width

                st.download_button(
                    label="📥 Download Audit Report",
                    data=output.getvalue(),
                    file_name=f"Audit_Report_{start_date.strftime('%d_%b')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                st.success("✓ Report ready! The file includes Summary, Exceptions, and Remarks sheets.")
        else:
            st.error("Invalid Date Format. Please use 'Day_Month' (e.g., 23_Jan or 30_Jan)")
    else:
        st.warning("⚠️ Please upload both files and enter a start date.")

# --- FOOTER ---
st.markdown("---")
st.markdown("**RDM Timesheet Auditor v2.0** | Enhanced with automatic time extraction and remarks tracking")
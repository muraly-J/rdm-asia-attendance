import pandas as pd
import openpyxl
from datetime import datetime, timedelta
import os
import re

def parse_date_from_filename(filename):
    """
    Attempts to extract start date from filename. 
    If not found, asks user.
    """
    pattern = r'(\d+)[_ ](\w+)'
    match = re.search(pattern, filename)
    if not match:
        user_input = input("Enter start date (e.g., 23_Jan): ").strip()
        return parse_user_date(user_input)
    
    start_day = int(match.group(1))
    start_month = match.group(2).capitalize()[:3]
    return finalize_date(start_day, start_month)

def parse_user_date(user_input):
    """Parses user input string into a date object."""
    try:
        # Expects formats like '23_Jan' or '23 Jan'
        clean_input = re.split(r'[_ ]', user_input)
        day = int(clean_input[0])
        month = clean_input[1].capitalize()[:3]
        return finalize_date(day, month)
    except:
        print("Invalid format. Use 'Day_Month' (e.g., 23_Jan)")
        return parse_user_date(input("Enter start date: "))

def finalize_date(day, month_str):
    current_year = 2026 
    month_map = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
                 'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12}
    return datetime(current_year, month_map.get(month_str, 1), day)

def clean_name(name):
    if pd.isna(name): return ""
    return re.sub(r'[^a-zA-Z0-9]', '', str(name).upper())

def extract_timesheet_data(timesheet_file, attendance_file):
    # Dynamic Date Setup
    start_date = parse_date_from_filename(timesheet_file)
    end_date = start_date + timedelta(days=6)
    date_range = [(start_date + timedelta(days=i)).date() for i in range(7)]
    
    # Generate the dynamic sheet name to look for (e.g., '23 Jan - 29 Jan')
    dynamic_sheet_name = f"{start_date.day} {start_date.strftime('%b')} - {end_date.day} {end_date.strftime('%b')}"

    if timesheet_file.endswith('.csv'):
        df = pd.read_csv(timesheet_file)
    else:
        df = pd.read_excel(timesheet_file)
    
    df['DateTime'] = pd.to_datetime(
        df['Date Time'].str.extract(r'(\d{4}-\d{2}-\d{2})')[0] + ' ' + 
        df['Date Time'].str.extract(r'(\d{2}:\d{2})')[0]
    )
    
    # Using 8-hour threshold as per your latest preference
    df['WorkDate'] = df.apply(lambda r: (r['DateTime'] - timedelta(hours=8)).date(), axis=1)
    df['TimeOnly'] = df['DateTime'].dt.time
    df['CleanName'] = df['Name'].apply(clean_name)

    # Dynamic Sheet Selection
    try:
        df_attendance = pd.read_excel(attendance_file, sheet_name=dynamic_sheet_name, header=None)
    except:
        print(f"⚠️ Could not find sheet '{dynamic_sheet_name}'. Using the first sheet instead.")
        df_attendance = pd.read_excel(attendance_file, sheet_name=0, header=None)
    
    raw_employee_list = [df_attendance.iloc[idx, 1] for idx in range(2, min(100, len(df_attendance))) 
                         if pd.notna(df_attendance.iloc[idx, 1])]
    attendance_map = {clean_name(name): name for name in raw_employee_list}
    
    output_data = []
    exception_logs = []

    for cleaned_name, original_name in attendance_map.items():
        row_times = []
        emp_records = df[df['CleanName'] == cleaned_name].sort_values('DateTime')

        for work_date in date_range:
            day_logs = emp_records[emp_records['WorkDate'] == work_date]
            login_time, logout_time = None, None
            
            if not day_logs.empty:
                # Login Priority
                start_work_logs = day_logs[day_logs['Type'] == 'Start Work']
                site_in_logs = day_logs[day_logs['Type'] == 'Site In']
                if not start_work_logs.empty:
                    login_time = start_work_logs.iloc[0]['TimeOnly']
                elif not site_in_logs.empty:
                    login_time = site_in_logs.iloc[0]['TimeOnly']
                else:
                    login_time = day_logs.iloc[0]['TimeOnly']

                # Logout Priority
                end_work_logs = day_logs[day_logs['Type'] == 'End Work']
                site_out_logs = day_logs[day_logs['Type'] == 'Site Out']
                if not end_work_logs.empty:
                    logout_time = end_work_logs.iloc[-1]['TimeOnly']
                elif not site_out_logs.empty:
                    logout_time = site_out_logs.iloc[-1]['TimeOnly']
                else:
                    logout_time = "NO LOG"

                if len(day_logs) == 1 and logout_time != "NO LOG":
                    if any(x in str(day_logs.iloc[0]['Type']) for x in ['Start Work', 'Site In']):
                        logout_time = "NO LOG"

                # Exception Logging
                if logout_time == "NO LOG":
                    exception_logs.append({'Name': original_name, 'Date': work_date, 'Time': login_time, 'Reason': 'Missing Logout'})
                elif "Site In" in day_logs['Type'].values and "Site Out" not in day_logs['Type'].values:
                    exception_logs.append({'Name': original_name, 'Date': work_date, 'Time': logout_time, 'Reason': 'Missing Site Out'})

            row_times.extend([login_time, logout_time])
        output_data.append(row_times)

    return pd.DataFrame(output_data), pd.DataFrame(exception_logs), date_range, raw_employee_list

def save_to_excel(df, date_range, output_file, df_exceptions, emp_list):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    bold, center, red_text = openpyxl.styles.Font(bold=True), openpyxl.styles.Alignment(horizontal='center'), openpyxl.styles.Font(color="FF0000", bold=True)

    ws.cell(1, 1, "No").font = bold
    ws.cell(1, 2, "Employee Name").font = bold
    
    for i, d_obj in enumerate(date_range):
        col = 3 + (i * 2)
        ws.cell(1, col, d_obj.strftime('%d %b')).font = bold
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)
        ws.cell(1, col).alignment = center
        ws.cell(2, col, "Login").font = bold
        ws.cell(2, col+1, "Logout").font = bold

    for r_idx, (emp_name, row_vals) in enumerate(zip(emp_list, df.itertuples(index=False)), start=3):
        ws.cell(r_idx, 1, r_idx - 2)
        ws.cell(r_idx, 2, emp_name)
        for c_idx, val in enumerate(row_vals, start=3):
            cell = ws.cell(r_idx, c_idx, val)
            if val == "NO LOG": cell.font = red_text
            elif val and not isinstance(val, str): cell.number_format = 'HH:MM'

    ws_ex = wb.create_sheet("Exception Audit")
    for i, h in enumerate(['Name', 'Date', 'Time', 'Reason'], 1): ws_ex.cell(1, i, h).font = bold
    for r_idx, row in df_exceptions.iterrows():
        ws_ex.cell(r_idx+2, 1, row['Name']); ws_ex.cell(r_idx+2, 2, row['Date'])
        ws_ex.cell(r_idx+2, 3, row['Time']).number_format = 'HH:MM'; ws_ex.cell(r_idx+2, 4, row['Reason'])

    wb.save(output_file)

def main():
    ts_file = input("Enter timesheet filename: ").strip()
    att_file = 'Attendance_New.xlsx'
    if not os.path.exists(ts_file): return print("File not found")
    
    try:
        df_out, df_ex, d_range, names = extract_timesheet_data(ts_file, att_file)
        out_name = f"Audit_Report_{d_range[0].strftime('%d_%b')}.xlsx"
        save_to_excel(df_out, d_range, out_name, df_ex, names)
        print(f"✓ Success! Created {out_name}")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    main()